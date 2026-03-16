import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def save_excel_with_autofit(df, filepath, index=False):
    """Save DataFrame to Excel with auto-fitted column widths."""
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=index, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']

        # Column offset (1 if index is written, 0 otherwise)
        col_offset = 1 if index else 0

        # Handle index column if present
        if index:
            max_idx_len = max(df.index.astype(str).map(
                len).max(), len(str(df.index.name) or '')) + 2
            worksheet.column_dimensions['A'].width = max_idx_len

        # Left-justify header cells
        header_row = 1
        for col_idx in range(1, len(df.columns) + 1 + col_offset):
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.alignment = Alignment(horizontal='left')

        for idx, col in enumerate(df.columns):
            # Get max length of column content
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(str(col))
            ) + 2  # Add padding
            worksheet.column_dimensions[get_column_letter(
                idx + 1 + col_offset)].width = max_length
